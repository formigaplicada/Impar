"""
gerar_fichas.py
---------------
Para cada loja com OneDrive configurado:
  1. Vai buscar o template (ficheiro que começa com "Template_Ficha") à pasta da loja
  2. Lista as subpastas (cada uma começa com o n_impar do condomínio)
  3. Para cada subpasta, vai à BD buscar os dados do condomínio
  4. Preenche as 8 células na sheet "Informacao do Condomínio"
  5. Insere o QR code como imagem na sheet "Ficha Condomínio"
  6. Grava "Ficha_{n_impar}.xlsx" na pasta do condomínio
  7. Se já existir, não faz nada (a menos que --force seja usado)

Uso:
  python gerar_fichas.py              -- processa todas as lojas
  python gerar_fichas.py --loja 9    -- processa só a loja com id=9
  python gerar_fichas.py --loja 9 --force  -- força regeneração mesmo que já exista
"""

import os
import io
import re
import sys
import requests
import psycopg2
import qrcode
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from msal import ConfidentialClientApplication
from dotenv import load_dotenv

load_dotenv()

TENANT_ID     = os.environ["MICROSOFT_TENANT_ID"]
CLIENT_ID     = os.environ["MICROSOFT_CLIENT_ID"]
CLIENT_SECRET = os.environ["MICROSOFT_CLIENT_SECRET"]
DATABASE_URL  = os.environ["DATABASE_URL"]

GRAPH_BASE    = "https://graph.microsoft.com/v1.0"
ONEDRIVE_USER = "vitor.lopes@impar.pt"

QR_W_PX = 157  # 1_500_000 EMU / 9525
QR_H_PX = 157

loja_filtro = None
if "--loja" in sys.argv:
    idx = sys.argv.index("--loja")
    loja_filtro = int(sys.argv[idx + 1])

force = "--force" in sys.argv


def get_token():
    app = ConfidentialClientApplication(
        CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{TENANT_ID}",
        client_credential=CLIENT_SECRET,
    )
    result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in result:
        raise Exception(f"Erro auth: {result.get('error_description')}")
    return result["access_token"]


def graph_get(token, path, params=None):
    r = requests.get(
        f"{GRAPH_BASE}{path}",
        headers={"Authorization": f"Bearer {token}"},
        params=params
    )
    r.raise_for_status()
    return r.json()


def graph_get_bytes(token, path):
    r = requests.get(
        f"{GRAPH_BASE}{path}",
        headers={"Authorization": f"Bearer {token}"}
    )
    r.raise_for_status()
    return r.content


def graph_upload(token, drive_id, folder_id, filename, content):
    url = f"{GRAPH_BASE}/drives/{drive_id}/items/{folder_id}:/{filename}:/content"
    r = requests.put(url, headers={
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }, data=content)
    r.raise_for_status()
    return r.json()


def get_lojas(conn, loja_filtro=None):
    with conn.cursor() as cur:
        if loja_filtro:
            cur.execute("""
                SELECT id, nome, onedrive_activos_folder_id
                FROM lojas
                WHERE ativo = true AND onedrive_activos_folder_id IS NOT NULL
                AND id = %s
            """, (loja_filtro,))
        else:
            cur.execute("""
                SELECT id, nome, onedrive_activos_folder_id
                FROM lojas
                WHERE ativo = true AND onedrive_activos_folder_id IS NOT NULL
                ORDER BY nome ASC
            """)
        cols = [d[0] for d in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]


def get_condominio(conn, n_impar):
    with conn.cursor() as cur:
        cur.execute("""
            SELECT n_impar, gestor, telefone2, email_gestor,
                   morada, codigo_postal, cidade, nipc
            FROM condominios
            WHERE n_impar = %s AND ativo = true
        """, (n_impar,))
        cols = [d[0] for d in cur.description]
        rows = cur.fetchall()
        if not rows:
            return None
        return dict(zip(cols, rows[0]))


def gerar_qr_image(nipc):
    url = f"https://my.condexpress.com/?condominio={nipc}"
    qr = qrcode.QRCode(box_size=10, border=2)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#011640", back_color="white").convert("RGB")
    img = img.resize((QR_W_PX, QR_H_PX), PILImage.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    xl = XLImage(buf)
    xl.width  = QR_W_PX
    xl.height = QR_H_PX
    return xl


def preencher_template(template_bytes, d):
    buf = io.BytesIO(template_bytes)
    wb = load_workbook(buf)

    # Preencher dados na sheet Informacao do Condominio
    ws_info = wb["Informacao do Condomínio"]
    ws_info["B2"]  = d.get("gestor")        or ""
    ws_info["B3"]  = d.get("telefone2")     or ""
    ws_info["B4"]  = d.get("email_gestor")  or ""
    ws_info["B5"]  = d.get("morada")        or ""
    ws_info["B6"]  = d.get("codigo_postal") or ""
    ws_info["B7"]  = d.get("cidade")        or ""
    ws_info["B10"] = d.get("nipc")          or ""
    ws_info["B13"] = d.get("n_impar")       or ""

    # Inserir QR na sheet Ficha Condominio (se tiver NIPC)
    if d.get("nipc") and "Ficha Condomínio" in wb.sheetnames:
        ws_ficha = wb["Ficha Condomínio"]

        # Remover imagem de QR existente (a que tem a formula IMAGE)
        # Manter todas as outras imagens (logo, fotos)
        ws_ficha._images = [img for img in ws_ficha._images
                            if not (hasattr(img.anchor, '_from') and
                                    img.anchor._from.col == 6 and
                                    img.anchor._from.row == 27)]

        # Inserir QR novo como imagem estática
        qr_img = gerar_qr_image(d["nipc"])
        qr_img.anchor = "G28"
        ws_ficha.add_image(qr_img)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def main():
    print("A autenticar no Microsoft...")
    token = get_token()

    print("A ligar à BD...")
    conn = psycopg2.connect(DATABASE_URL)

    drive_info = graph_get(token, f"/users/{ONEDRIVE_USER}/drive")
    drive_id = drive_info["id"]
    print(f"Drive ID: {drive_id}")

    lojas = get_lojas(conn, loja_filtro)
    print(f"\n{len(lojas)} loja(s) a processar\n")

    total_criadas   = 0
    total_ignoradas = 0
    total_erros     = 0

    for loja in lojas:
        folder_id = loja["onedrive_activos_folder_id"]
        print(f"\n-- Loja: {loja['nome']}")

        try:
            items = graph_get(token, f"/drives/{drive_id}/items/{folder_id}/children", {"$top": 500})
            filhos = items.get("value", [])
        except Exception as e:
            print(f"   ERRO ao listar pasta: {e}")
            continue

        template_item = next(
            (f for f in filhos if f.get("name", "").startswith("Template_Ficha") and f["name"].endswith(".xlsx")),
            None
        )
        if not template_item:
            print(f"   AVISO: Sem template encontrado")
            continue

        print(f"   Template: {template_item['name']}")

        try:
            template_bytes = graph_get_bytes(token, f"/drives/{drive_id}/items/{template_item['id']}/content")
        except Exception as e:
            print(f"   ERRO ao descarregar template: {e}")
            continue

        subpastas = [f for f in filhos if f.get("folder") and not f["name"].startswith("Template")]
        print(f"   {len(subpastas)} subpasta(s)")

        for pasta in subpastas:
            nome_pasta = pasta["name"]

            match = re.match(r'^(\d+)', nome_pasta)
            if not match:
                print(f"   SKIP '{nome_pasta}' -- sem n_impar")
                continue

            n_impar = int(match.group(1))

            if n_impar >= 100000:
                total_ignoradas += 1
                continue

            pasta_id      = pasta["id"]
            nome_ficheiro = f"Ficha_{n_impar}.xlsx"

            try:
                filhos_pasta = graph_get(token, f"/drives/{drive_id}/items/{pasta_id}/children")
                ja_existe = any(f["name"] == nome_ficheiro for f in filhos_pasta.get("value", []))
            except Exception as e:
                print(f"   ERRO ao verificar '{nome_pasta}': {e}")
                total_erros += 1
                continue

            if ja_existe and not force:
                print(f"   SKIP {n_impar} -- ficha ja existe")
                total_ignoradas += 1
                continue

            d = get_condominio(conn, n_impar)
            if not d:
                print(f"   SKIP {n_impar} -- nao encontrado na BD")
                total_ignoradas += 1
                continue

            try:
                ficha_bytes = preencher_template(template_bytes, d)
                try:
                    graph_upload(token, drive_id, pasta_id, nome_ficheiro, ficha_bytes)
                    print(f"   OK {n_impar}")
                except requests.exceptions.HTTPError as e:
                    if force and e.response is not None and e.response.status_code == 423:
                        nome_fallback = nome_ficheiro.replace(".xlsx", "_1.xlsx")
                        graph_upload(token, drive_id, pasta_id, nome_fallback, ficha_bytes)
                        print(f"   OK {n_impar} (locked, gravado como {nome_fallback})")
                    else:
                        raise
                total_criadas += 1
            except Exception as e:
                print(f"   ERRO {n_impar}: {e}")
                total_erros += 1

    conn.close()
    print(f"\n{'='*50}")
    print(f"Concluido: {total_criadas} criadas | {total_ignoradas} ignoradas | {total_erros} erros")


if __name__ == "__main__":
    main()
