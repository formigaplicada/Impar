import os
import io
import shutil
import subprocess
import tempfile
import urllib.request
from fastapi import FastAPI, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import qrcode
from PIL import Image as PILImage

app = FastAPI()

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "template.xlsx")

# Dimensões em EMUs (English Metric Units)
# 1 inch = 914400 EMUs, 1 px (96dpi) = 9525 EMUs
# Fotos da loja: ocupam metade da largura útil da página (~9cm cada)
FOTO_W_EMU = 2_750_000
FOTO_H_EMU = 1_800_000
QR_W_EMU   = 1_500_000
QR_H_EMU   = 1_500_000

class DadosCondominio(BaseModel):
    nome:          str | None = None
    morada:        str | None = None
    codigo_postal: str | None = None
    cidade:        str | None = None
    nipc:          str | None = None
    n_impar:       int | None = None
    loja_nome:     str | None = None
    loja_morada:   str | None = None
    loja_email:    str | None = None
    loja_telefone: str | None = None
    loja_foto1:    str | None = None
    loja_foto2:    str | None = None
    gestor_nome:   str | None = None
    gestor_email:  str | None = None
    gestor_tel:    str | None = None

def download_image(url: str) -> PILImage.Image | None:
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as r:
            return PILImage.open(io.BytesIO(r.read())).convert("RGB")
    except Exception:
        return None

def pil_to_xl_image(pil_img: PILImage.Image, w_emu: int, h_emu: int) -> XLImage:
    w_px = w_emu // 9525
    h_px = h_emu // 9525
    # Crop centralizado para forçar ratio exato
    src_w, src_h = pil_img.size
    src_ratio = src_w / src_h
    dst_ratio = w_px / h_px
    if src_ratio > dst_ratio:
        new_w = int(src_h * dst_ratio)
        left = (src_w - new_w) // 2
        pil_img = pil_img.crop((left, 0, left + new_w, src_h))
    else:
        new_h = int(src_w / dst_ratio)
        top = (src_h - new_h) // 2
        pil_img = pil_img.crop((0, top, src_w, top + new_h))
    pil_img = pil_img.resize((w_px, h_px), PILImage.LANCZOS)
    buf = io.BytesIO()
    pil_img.save(buf, format="PNG")
    buf.seek(0)
    xl = XLImage(buf)
    xl.width  = w_px
    xl.height = h_px
    return xl

def gerar_qr(url: str) -> XLImage:
    qr = qrcode.QRCode(box_size=10, border=2)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#011640", back_color="white").convert("RGB")
    w_px = QR_W_EMU // 9525
    h_px = QR_H_EMU // 9525
    img = img.resize((w_px, h_px), PILImage.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    xl = XLImage(buf)
    xl.width  = w_px
    xl.height = h_px
    return xl
@app.get("/health")
def health():
    return {"ok": True}

@app.post("/gerar-ficha")
def gerar_ficha(dados: DadosCondominio):
    tmpdir = tempfile.mkdtemp()
    try:
        xlsx_path = os.path.join(tmpdir, "ficha.xlsx")
        shutil.copy(TEMPLATE_PATH, xlsx_path)

        wb = load_workbook(xlsx_path)
        ws = wb["Ficha Condomínio"]

        # Dados do condomínio
        ws["C6"] = dados.morada or ""
        ws["B7"] = dados.codigo_postal or ""
        ws["J7"] = dados.cidade or ""

        # Dados da loja
        ws["C16"] = dados.loja_nome or ""
        ws["C17"] = dados.loja_morada or ""
        loja_contacto = ""
        if dados.loja_email:
            loja_contacto += dados.loja_email
        if dados.loja_telefone:
            loja_contacto += f"  |  {dados.loja_telefone}"
        ws["C18"] = loja_contacto

        # Dados do gestor
        ws["C23"] = dados.gestor_nome or ""
        gestor_contacto = ""
        if dados.gestor_email:
            gestor_contacto += dados.gestor_email
        if dados.gestor_tel:
            gestor_contacto += f"  |  {dados.gestor_tel}"
        ws["C24"] = gestor_contacto

        # Limpar célula do QR (tinha fórmula #REF!)
        ws["G28"] = ""

        # Remover fotos antigas (manter só o logo — img0)
        logo = ws._images[0] if ws._images else None
        ws._images = []
        if logo:
            ws._images.append(logo)

        # Inserir fotos da loja
        if dados.loja_foto1:
            img = download_image(dados.loja_foto1)
            if img:
                xl = pil_to_xl_image(img, FOTO_W_EMU, FOTO_H_EMU)
                xl.anchor = "C10"
                ws.add_image(xl)

        if dados.loja_foto2:
            img = download_image(dados.loja_foto2)
            if img:
                xl = pil_to_xl_image(img, FOTO_W_EMU, FOTO_H_EMU)
                xl.anchor = "J10"
                ws.add_image(xl)

        # QR code
        if dados.nipc:
            qr = gerar_qr(f"https://my.condexpress.com/?condominio={dados.nipc}")
            qr.anchor = "G28"
            ws.add_image(qr)

        wb.save(xlsx_path)

        # Converter para PDF
        result = subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "pdf",
             "--outdir", tmpdir, xlsx_path],
            capture_output=True, text=True, timeout=60
        )

        if result.returncode != 0:
            raise HTTPException(status_code=500, detail=f"LibreOffice error: {result.stderr}")

        pdf_path = os.path.join(tmpdir, "ficha.pdf")
        if not os.path.exists(pdf_path):
            raise HTTPException(status_code=500, detail="PDF não foi gerado")

        nome = f"ficha_{dados.n_impar or 'condominio'}.pdf"
        with open(pdf_path, "rb") as f:
            pdf_bytes = f.read()

        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{nome}"'}
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)
